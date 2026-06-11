from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple


@dataclass(frozen=True)
class LoadedContext:
    used_files: List[str]
    context: str
    context_note: Optional[str]
    file_errors: List[str]


@dataclass(frozen=True)
class IndexedFile:
    file: str
    content: str


@dataclass(frozen=True)
class IndexingInput:
    files: List[Dict[str, str]]
    notes: List[str]
    file_errors: List[str]


SUPPORTED_EXTS: Set[str] = {".txt", ".md", ".json", ".csv", ".pdf", ".docx", ".xlsx"}


def _resolve_path(project_root: Path, p: str) -> Path:
    path = Path(p)
    if path.is_absolute():
        return path
    return (project_root / path).resolve()


def _read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def read_pdf_text(
    path: Path,
    max_chars: int,
    *,
    include_page_markers: bool = False,
    source_file_label: str = "",
) -> str:
    """Extract plain text from PDF via PyMuPDF (no markdown conversion)."""
    import fitz  # PyMuPDF

    limit = max_chars if max_chars > 0 else 10**9
    doc = fitz.open(str(path))
    try:
        parts: List[str] = []
        total = 0
        for idx, page in enumerate(doc, start=1):
            txt = (page.get_text("text") or "").strip()
            if not txt:
                continue
            if include_page_markers and source_file_label:
                parts.append(f"[[SOURCE file={source_file_label} page={idx}]]")
            parts.append(txt)
            total += len(txt)
            if total >= limit:
                break
        return "\n".join(parts).strip()
    finally:
        doc.close()


def _read_pdf(
    path: Path,
    max_chars_hint: int,
    *,
    include_page_markers: bool,
    source_file_label: str,
) -> str:
    return read_pdf_text(
        path,
        max_chars_hint,
        include_page_markers=include_page_markers,
        source_file_label=source_file_label,
    )


def _read_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    paras = [p.text for p in doc.paragraphs if p.text]
    return "\n".join(paras)


def _read_xlsx(path: Path, *, max_sheets: int = 5, max_rows: int = 200, max_cols: int = 30) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames[:max_sheets]
    out: List[str] = []
    for name in sheet_names:
        ws = wb[name]
        out.append(f"## Sheet: {name}")
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx > max_rows:
                out.append("...(more rows truncated)")
                break
            values = []
            for c_idx, v in enumerate(row, start=1):
                if c_idx > max_cols:
                    values.append("...(more cols truncated)")
                    break
                if v is None:
                    values.append("")
                else:
                    values.append(str(v))
            out.append("\t".join(values).rstrip())
    return "\n".join(out)


def load_agent_files(*, project_root: Path, file_paths: List[str], max_file_chars: int) -> LoadedContext:
    used_files: List[str] = []
    file_errors: List[str] = []
    chunks: List[str] = []

    for raw in file_paths:
        p = _resolve_path(project_root, raw)
        used_files.append(raw)
        try:
            if not p.exists():
                file_errors.append(f"文件不存在：{raw}")
                continue

            ext = p.suffix.lower()
            if ext in {".txt", ".md", ".json", ".csv"}:
                content = _read_text_file(p)
            elif ext == ".pdf":
                content = _read_pdf(
                    p,
                    max_chars_hint=max_file_chars,
                    include_page_markers=True,
                    source_file_label=raw,
                )
            elif ext == ".docx":
                content = _read_docx(p)
            elif ext == ".xlsx":
                content = _read_xlsx(p)
            else:
                file_errors.append(f"暂不支持的文件类型：{raw}")
                continue

            content = (content or "").strip()
            if not content:
                file_errors.append(f"文件为空或无可提取文本：{raw}")
                continue

            # Add a file-level source marker so we can later attribute snippets.
            if ext != ".pdf":
                content = f"[[SOURCE file={raw}]]\n{content}"
            chunks.append(f"===== 文件：{raw} =====\n{content}")
        except Exception as e:  # noqa: BLE001
            file_errors.append(f"读取失败：{raw}（{type(e).__name__}: {e}）")

    context = "\n\n".join(chunks).strip()
    context_note: Optional[str] = None
    if context and len(context) > max_file_chars:
        context = context[:max_file_chars]
        context_note = f"文件内容较长，已读取前 {max_file_chars} 字符进行分析"

    return LoadedContext(
        used_files=used_files,
        context=context,
        context_note=context_note,
        file_errors=file_errors,
    )


def list_supported_files_in_dir(*, project_root: Path, files_dir: str, recursive: bool = False) -> Tuple[List[str], List[str]]:
    """
    Return (file_paths, errors). Paths are normalized to posix style; relative to project_root when possible.
    Non-recursive by default.
    """
    errors: List[str] = []
    if not files_dir:
        return [], errors

    dir_path = _resolve_path(project_root, files_dir)
    if not dir_path.exists():
        errors.append(f"目录不存在：{files_dir}")
        return [], errors
    if not dir_path.is_dir():
        errors.append(f"不是目录：{files_dir}")
        return [], errors

    if recursive:
        candidates = [p for p in dir_path.rglob("*") if p.is_file()]
    else:
        candidates = [p for p in dir_path.iterdir() if p.is_file()]

    out: List[str] = []
    for p in sorted(candidates, key=lambda x: x.name.lower()):
        if p.suffix.lower() not in SUPPORTED_EXTS:
            continue
        try:
            rel = p.relative_to(project_root).as_posix()
        except Exception:
            rel = str(p)
        out.append(rel)
    return out, errors


def expand_agent_file_list(*, project_root: Path, files: List[str], files_dir: str) -> Tuple[List[str], List[str]]:
    """
    Expand a single agent's file inputs into an ordered, de-duplicated file list.
    - First include dir-scanned files (non-recursive)
    - Then include explicitly registered files
    """
    dir_files, dir_errors = list_supported_files_in_dir(project_root=project_root, files_dir=files_dir, recursive=False)
    merged: List[str] = []
    seen = set()
    for f in dir_files + list(files or []):
        if not isinstance(f, str):
            continue
        f = f.strip()
        if not f or f in seen:
            continue
        merged.append(f)
        seen.add(f)
    return merged, dir_errors


def load_files_for_indexing(*, project_root: Path, file_paths: List[str], max_file_chars: int) -> IndexingInput:
    files: List[Dict[str, str]] = []
    notes: List[str] = []
    file_errors: List[str] = []

    for raw in file_paths:
        p = _resolve_path(project_root, raw)
        try:
            if not p.exists():
                file_errors.append(f"文件不存在：{raw}")
                continue

            ext = p.suffix.lower()
            if ext in {".txt", ".md", ".json", ".csv"}:
                content = _read_text_file(p)
            elif ext == ".pdf":
                content = _read_pdf(
                    p,
                    max_chars_hint=max_file_chars,
                    include_page_markers=False,
                    source_file_label=raw,
                )
            elif ext == ".docx":
                content = _read_docx(p)
            elif ext == ".xlsx":
                content = _read_xlsx(p)
            else:
                file_errors.append(f"暂不支持的文件类型：{raw}")
                continue

            content = (content or "").strip()
            if not content:
                file_errors.append(f"文件为空或无可提取文本：{raw}")
                continue

            if len(content) > max_file_chars:
                content = content[:max_file_chars]
                notes.append(f"{raw} 内容较长，已读取前 {max_file_chars} 字符用于生成索引")

            files.append({"file": raw, "content": content})
        except Exception as e:  # noqa: BLE001
            file_errors.append(f"读取失败：{raw}（{type(e).__name__}: {e}）")

    return IndexingInput(files=files, notes=notes, file_errors=file_errors)

